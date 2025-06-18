# --Requirements--
# install OpenPyXL  -- pip install OpenPyXL      = for reading purpose
# install xlsxwriter -- pip install xlsxwriter   => for writing purpose

import xlsxwriter #importing xlsxwriter
import openpyxl #importing openpyxl
from openpyxl import Workbook, load_workbook   #helps in reading the excel file


def hashing(password):  #Function which do the hashing and take a parameter "password"
    #dict1 stores most of the keys as key in dict and its hash as value
    dict1 ={
    "a": 'a1', "b": 'b2', "c": "c3", "d": "d4", "e": "e5", "f": "f6", "g": "g7", "h": "h8", "i": "i9", "j": "j10", 
    "k": "k11", "l": "l12", "m": "m13", "n": "n14", "o": "o15", "p": "p16", "q": "q17", "r": "r18", "s": "s19",
    "t": "t20", "u": "u21", "v": "v22", "w": "w23", "x": "x24", "y": "y25", "z": "z26", "A": "A1", "B": "B2",
    "C": "C3", "D": "D4", "E": "E5", "F": "F6", "G": "G7", "H": "H8", "I": "I9", "J": "J10", "K": "K11", "L": "L12",
    "M": "M13", "N": "N14", "O": "O15", "P": "P16", "Q": "Q17", "R": "R18", "S": "S19", "T": "T20", "U": "U21",
    "V": "V22", "W": "W23", "X": "X24", "Y": "Y25", "Z": "Z26", "0": "01", "1": "12", "2": "23", "3": "34",
    "4": "45", "5": "56", "6": "67", "7": "78", "8": "89", "9": "910", "!": "!1", "@": "@2", "#": "#3", "$": "$4",
    "%": "%5", "^": "^6", "&": "&7", "*": "*8", "(": "(9", ")": ")10", "-": "-11", "_": "_12", "=": "=13",
    "+": "+14", "{": "{15", "}": "}16", "[": "[17", "]": "]18", "\\": "\\19", "|": "|20", ";": ";21", ":": ":22",
    "'": "'23", '"': '"24', ",": ",25", ".": ".26", "/": "/27", "?": "?28", "<": "<29", ">": ">30", "~": "~31",
    "`": "`32", "§": "§33", "±": "±34", "÷": "÷35", "*": "*36", "®": "®37", "©": "©38", "™": "™39","€": "€40",
    "£": "£41", "¥": "¥42", "¢": "¢43", "¡": "¡44", "¿": "¿45", "¶": "¶46", "†": "†47", "‡": "‡48", "•": "•49",
    " ": " 50"
    }
    

    list1 = [*password]  #seprate the password - from word to individual letters
    hashed_password = ''  #empty string
    for i in list1:  #loop checks the each value of list1 in dict1 for hashing
        hashed_password += str(dict1[i]) #stores the hash in hashed_password

    return hashed_password #returns hashed_password


def WriteExcel(): #Function which opens the excel file and do the entry part
    workbook = xlsxwriter.Workbook("UserData.xlsx") #creates a excel file which is called workbook
    worksheet = workbook.add_worksheet("DataSheet1")  #adds a worksheet in the workbook
    worksheet.write(0, 0, "S_No")  #adds a name to the coloumn in the excel file
    worksheet.write(0, 1, "Name")
    worksheet.write(0, 2, "Age")
    worksheet.write(0, 3, "Gender")
    worksheet.write(0, 4, "DOB")
    worksheet.write(0, 5, "Address")
    worksheet.write(0, 6, "Phone_No")
    worksheet.write(0, 7, "Email_ID")
    worksheet.write(0, 8, "Org_Password")
    worksheet.write(0, 9, "Hash_Password")

    workbook.close()   #saves the created excel file

    print("\nCreated Data file Successfully\n")


def ReadExcel(): #Function which opens the excel file and do the reading part
    n = int(input("\nEnter no of entries to add: "))  # 'n' helps in ending the loop
    m = int(input("Enter no of the last row entry in data file: ")) # 'm' helps in finding the start of loop

    # m+1 sets the cursor in the new row as m denotes the previous filled row and m+1 denotes new row.
    # m+1+n stops the cursor in the mentioned row as n denotes the no. of entries to add and they will be
    # counted after the fresh row i.e. m+1

    for i in range(m+1, m+1+n):
        book = load_workbook("UserData.xlsx") #opems the excel file
        #allow user to use the active sheet of the excel file i.e the first sheet in woorkbook (excel file)
        sheet = book.active

        print("\nPLease provide the details")
        #input from the user to get his/her details 
        serial = int(input("Enter Serial Number: "))
        name = input("Enter user's name: ")
        age = int(input("Enter user's age: "))
        gender = input("Enter user's gender: ")
        dob = input("Enter user's data of birth: ")
        address = input("Enter user's address: ")
        phone_no = input("Enter user's phone number: ")
        email_id = input("Enter user's email id: ")
        password = input("Enter password for ID: ")
        hashed_password = hashing(password) #calls the hashing function
        
        #modifies the data and add the details in the excel file as the loops continues.
        sheet[f'A{i}'].value = serial
        sheet[f'B{i}'].value = name
        sheet[f'C{i}'].value = age
        sheet[f'D{i}'].value = gender
        sheet[f'E{i}'].value = dob
        sheet[f'F{i}'].value = address
        sheet[f'G{i}'].value = phone_no
        sheet[f'H{i}'].value = email_id
        sheet[f'I{i}'].value = password
        sheet[f'J{i}'].value = hashed_password
        book.save("UserData.xlsx") #saves the change in excel file
    print("\nData entry completed successfully\n")


def main_menu(): #Function which displays the main menu
    ch = input("Press 1 to create an Excel data file \nPress 2 to add new entry to Excel File \nPress 0 to Quit \nPress Key: ")
    if ch == '1': #if condition is true
        WriteExcel() #calls WriteExcel function
        main() #calls the mains function

    elif ch == '2': #if condition is true
        ReadExcel() #calls ReadExcel function
        main() #calls the mains function

    elif ch == '0': #if condition is true
        print("\nThanks for visiting ComGames!! \nSee you again!\n")
        quit() #quits the program as the program is in loop untill and unless there is no input error.

    else: #if users enters the wrong input then else will be executed and again the main_menu will be called
        print("\nSorry! Not able to understand \n")
        main_menu() #calls main_menu function


def Greeting():  #Function which greets the user
    reply_1 = input("\nHey there, welcome to the ComGames! ")
    if reply_1 == "Hey" or reply_1 == "hey" or reply_1 == "HEY" :
        print ("\nHow can I help?")

    elif reply_1 == "Hi" or reply_1 == "hi" or reply_1 == "HI":
        print ("\nHow can I help?")

    elif reply_1 == "Hello" or reply_1 == "hello" or reply_1 == "HELLO":
        print ("\nHow can I help?")

    else:
        print ("Can't understand you....Sorry\n")
        quit()


def main():  #Function which stores the full program
    main_menu() #calling main_menu function to display menu from which the user will choose

    main() #re-calling main function like a loop


Greeting()  #calling greeting function
main() #calling main function

